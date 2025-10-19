from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment

from shared_code.application.app_directory_creator import AppDirectoryCreator
from shared_code.application.app_file_utils import AppFileUtils
from shared_code.domain.app_config import AppConfig
from shared_code.domain.table.table_name.set import TableNameSet
from shared_code.infra.database.mysql.entity_definitions_getter import (
    EntityDefinitionsGetRequest,
    EntityDefinitionsGetResponse,
    EntityDefinitionsGetter,
)
from shared_code.infra.file_system.app_config_jsonc_file_reader import (
    AppConfigJsoncFileReader,
)


class DMLSourceXlsxCreator:
    @classmethod
    def execute(
        cls,
        app_config_json_file_path_str: Optional[str],
        db_config_json_file_path_str: Optional[str],
        target_tables: Optional[str],
        xlsx_directory_path_str: Optional[str],
        dml_source_file_name: Optional[str],
    ) -> str:
        app_config_file_path_str = AppFileUtils.determine_and_check(
            arg_file_path_str=app_config_json_file_path_str,
            default_file_path=Path(__file__).parent.parent.parent.parent.joinpath(
                "app_config.json"
            ),
        )

        db_config_file_path_str = AppFileUtils.determine_and_check(
            arg_file_path_str=db_config_json_file_path_str,
            default_file_path=Path(__file__).parent.parent.parent.parent.joinpath(
                "db_config.json"
            ),
        )

        xlsx_dir_path_str = AppDirectoryCreator.execute(
            module_name="dml_source_xlsx",
            directory_type="xlsx output",
            directory_path_str=xlsx_directory_path_str,
        )

        xlsx_file_name = cls.__get_xlsx_file_name(file_name=dml_source_file_name)

        xlsx_file_path = Path(xlsx_dir_path_str).joinpath(xlsx_file_name)
        xlsx_file_path_str = str(xlsx_file_path)

        target_table_set = TableNameSet.from_csv_str(csv_str=target_tables)

        entity_definitions_get_request = EntityDefinitionsGetRequest(
            db_config_json_file_path_str=db_config_file_path_str,
            target_table_name_set=target_table_set,
        )
        entity_definitions_get_response = EntityDefinitionsGetter.execute(
            a_request=entity_definitions_get_request
        )

        app_config = AppConfigJsoncFileReader.execute(
            file_path=app_config_file_path_str
        )

        cls.__create_xlsx(
            app_config=app_config,
            entity_definitions_get_response=entity_definitions_get_response,
            xlsx_file_path_str=xlsx_file_path_str,
        )

        return xlsx_file_path_str

    @classmethod
    def __get_xlsx_file_name(cls, file_name: Optional[str]) -> str:
        if file_name:
            return file_name

        return "dml_source.xlsx"

    @classmethod
    def __create_xlsx(
        cls,
        app_config: AppConfig,
        entity_definitions_get_response: EntityDefinitionsGetResponse,
        xlsx_file_path_str: str,
    ):
        table_name_cell_position = app_config.table_name_cell_position
        data_start_cell_position = app_config.data_start_cell_position
        columns_row_header_column = data_start_cell_position.column - 1

        table_name_set = entity_definitions_get_response.table_name_set
        can_replace_sheet_name_with_table_name = (
            table_name_set.can_replace_xlsx_sheet_name()
        )

        a_workbook = Workbook()
        a_worksheet = a_workbook.active

        is_not_first_sheet = False
        for entity_definition in entity_definitions_get_response.entity_definitions.unmodifiable_elements.values():
            table_name = entity_definition.table_name
            table_name_value = table_name.value

            if is_not_first_sheet:
                a_worksheet = a_workbook.create_sheet()

            is_not_first_sheet = True

            if can_replace_sheet_name_with_table_name:
                a_worksheet.title = table_name_value

            a_worksheet.cell(
                row=table_name_cell_position.row,
                column=table_name_cell_position.column,
                value=table_name_value,
            )

            a_worksheet.cell(
                row=app_config.db_column_name_row_number.value,
                column=columns_row_header_column,
                value="COLUMN_NAME",
            )

            a_worksheet.cell(
                row=app_config.db_column_comment_row_number.value,
                column=columns_row_header_column,
                value="COLUMN_COMMENT",
            )

            a_worksheet.cell(
                row=app_config.column_default_row_number.value,
                column=columns_row_header_column,
                value="COLUMN_DEFAULT",
            )

            a_worksheet.cell(
                row=app_config.nullable_column_flag_row_number.value,
                column=columns_row_header_column,
                value="IS_NULLABLE",
            )

            a_worksheet.cell(
                row=app_config.data_type_row_number.value,
                column=columns_row_header_column,
                value="DATA_TYPE",
            )

            a_worksheet.cell(
                row=app_config.key_position_row_number.value,
                column=columns_row_header_column,
                value="KEY_POSITION",
            )

            a_cell = a_worksheet.cell(
                row=app_config.no_quotation_row_number.value,
                column=columns_row_header_column,
                value="no_quotation",
            )
            a_cell.comment = Comment(
                text="If you do not add quotation marks, enter some value.",
                author="python_sql_tools",
            )

            current_column = data_start_cell_position.column
            for db_column in entity_definition.db_columns.unmodifiable_elements:
                a_worksheet.cell(
                    row=app_config.db_column_name_row_number.value,
                    column=current_column,
                    value=db_column.column_name.value,
                )

                column_comment = db_column.column_comment
                a_worksheet.cell(
                    row=app_config.db_column_comment_row_number.value,
                    column=current_column,
                    value=column_comment.value if column_comment else "",
                )

                column_default = db_column.column_default
                a_worksheet.cell(
                    row=app_config.column_default_row_number.value,
                    column=current_column,
                    value=column_default.value if column_default else "",
                )

                a_worksheet.cell(
                    row=app_config.nullable_column_flag_row_number.value,
                    column=current_column,
                    value=str(db_column.nullable_column_flag).upper(),
                )

                a_worksheet.cell(
                    row=app_config.data_type_row_number.value,
                    column=current_column,
                    value=db_column.data_type.value,
                )

                key_position = db_column.key_position
                a_worksheet.cell(
                    row=app_config.key_position_row_number.value,
                    column=current_column,
                    value=str(key_position.value) if key_position else "",
                )

                current_column += 1


        a_workbook.save(xlsx_file_path_str)
