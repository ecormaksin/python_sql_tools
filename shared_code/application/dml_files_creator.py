from dataclasses import dataclass

from openpyxl import load_workbook

from shared_code.domain.app_config import AppConfig
from shared_code.domain.sink_dml_dir_path import SinkDMLDirectoryPath
from shared_code.domain.source_data_xlsx_file_path import SourceDataXlsxFilePath
from shared_code.domain.table_name import TableName
from shared_code.domain.table_name_definition_type import TableNameDefinitionType


@dataclass(frozen=True)
class DMLFilesCreationRequest:
    source_data_xlsx_file_path: SourceDataXlsxFilePath
    sink_dml_dir_path: SinkDMLDirectoryPath
    app_config: AppConfig


class DMLFilesCreator:
    def __init__(self, a_request: DMLFilesCreationRequest):
        self.__a_request = a_request

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, a_traceback):
        """
        do nothing
        """

    def execute(self):
        src_excel_file_path = self.__a_request.source_data_xlsx_file_path
        app_config = self.__a_request.app_config

        a_workbook = load_workbook(
            filename=src_excel_file_path.value, data_only=True, read_only=True
        )

        sheet_name = "media_type"
        a_worksheet = a_workbook[sheet_name]

        table_name = ""
        if app_config.table_name_definition_type == TableNameDefinitionType.SHEET:
            table_name = TableName(value=sheet_name)

        header_range = [
            [str(cell) for cell in row]
            for row in a_worksheet.iter_rows(
                min_row=1,
                max_row=app_config.header_max_row,
                min_col=app_config.data_start_cell_position.column,
                values_only=True,
            )
        ]
        print(header_range)

        data_range = [
            [str(cell) for cell in row]
            for row in a_worksheet.iter_rows(
                min_row=app_config.data_start_cell_position.row,
                max_row=a_worksheet.max_row,
                min_col=app_config.data_start_cell_position.column,
                values_only=True,
            )
        ]
        print(data_range)

        # with DMLCreator(
        #     table_name=table_name, a_worksheet=a_worksheet, app_config=app_config
        # ) as a_dml_creator:
        #     dmls = a_dml_creator.execute()

        # for sheet_name in a_workbook.sheetnames:
        #     a_worksheet = a_workbook[sheet_name]
        #
        #     if app_config.table_name_definition_type == TableNameDefinitionType.SHEET:
        #         table_name = TableName(value=sheet_name)
        #
        #     print("")
        #     print("---")
        #     print(sheet_name)
        #     print("min_row: " + str(a_worksheet.min_row))
        #     print("max_row: " + str(a_worksheet.max_row))
        #     print("min_column: " + str(a_worksheet.min_column))
        #     print("max_column: " + str(a_worksheet.max_column))
