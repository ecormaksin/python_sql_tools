from dataclasses import dataclass

from openpyxl import load_workbook

from shared_code.application.db_column.list_builder import (
    DBColumnsBuilder,
    DBColumnsBuildRequest,
)
from shared_code.application.dml.db_columns_data_range_getter import (
    DBColumnsDataRangeGetRequest,
    DBColumnsDataRangeGetter,
)
from shared_code.application.dml.dml_data_range_getter import (
    DMLDataRangeGetRequest,
    DMLDataRangeGetter,
)
from shared_code.application.dml.one_table_dmls_builder import (
    OneTableDMLsBuilder,
    OneTableDMLsBuildRequest,
)
from shared_code.application.dml.table_name_getter import (
    TableNameGetRequest,
    TableNameGetter,
)
from shared_code.domain.app_config import AppConfig
from shared_code.domain.dmls.entity import OneTableDmls
from shared_code.domain.dmls.list import AllTableDMLs
from shared_code.domain.source_data_xlsx_file_path import SourceDataXlsxFilePath


@dataclass(frozen=True)
class AllTablesDMLsBuildRequest:
    source_data_xlsx_file_path: SourceDataXlsxFilePath
    app_config: AppConfig


class AllTablesDMLsBuilder:
    @classmethod
    def execute(cls, a_request: AllTablesDMLsBuildRequest) -> AllTableDMLs:
        src_excel_file_path = a_request.source_data_xlsx_file_path
        app_config = a_request.app_config
        data_start_cell_position = app_config.data_start_cell_position
        target_sheet_names = app_config.target_sheet_names
        exclude_sheet_names = app_config.exclude_sheet_names

        a_workbook = load_workbook(filename=src_excel_file_path.value, data_only=True)

        all_tables_dmls = AllTableDMLs.empty()
        for sheet_name in a_workbook.sheetnames:
            if target_sheet_names.not_contains(sheet_name=sheet_name):
                continue

            if exclude_sheet_names.contains(sheet_name=sheet_name):
                continue

            a_worksheet = a_workbook[sheet_name]

            table_name = TableNameGetter.execute(
                a_request=TableNameGetRequest(
                    a_worksheet=a_worksheet,
                    table_name_cell_position=app_config.table_name_cell_position,
                )
            )

            db_columns_range = DBColumnsDataRangeGetter.execute(
                a_request=DBColumnsDataRangeGetRequest(
                    a_worksheet=a_worksheet,
                    max_row=app_config.header_max_row,
                    min_col=data_start_cell_position.column,
                )
            )
            # print(db_columns_range)

            db_columns_build_request = DBColumnsBuildRequest(
                source_data=db_columns_range, app_config=app_config
            )
            db_columns = DBColumnsBuilder.execute(a_request=db_columns_build_request)

            data_range = DMLDataRangeGetter.execute(
                a_request=DMLDataRangeGetRequest(
                    a_worksheet=a_worksheet,
                    data_start_cell_position=data_start_cell_position,
                )
            )
            # print(data_range)

            dmls_build_request = OneTableDMLsBuildRequest(
                table_name=table_name,
                db_columns=db_columns,
                data_range=data_range,
                set_empty_str_instead_of_null=app_config.set_empty_str_instead_of_null,
            )
            dml_list = OneTableDMLsBuilder.execute(a_request=dmls_build_request)
            dmls = OneTableDmls(table_name=table_name, dmls=dml_list)

            all_tables_dmls = all_tables_dmls.append(element=dmls)

        return all_tables_dmls
