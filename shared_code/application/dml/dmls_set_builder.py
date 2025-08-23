from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from shared_code.application.db_column.list_builder import (
    DBColumnsBuildRequest,
    DBColumnsBuilder,
)
from shared_code.application.dml.dmls_builder import DMLsBuilder, DMLsBuildRequest
from shared_code.domain.app_config import AppConfig
from shared_code.domain.dmls.entity import DMLs
from shared_code.domain.dmls.set import DMLsSet
from shared_code.domain.source_data_xlsx_file_path import SourceDataXlsxFilePath
from shared_code.domain.table_name import TableName
from shared_code.domain.table_name_definition_type import TableNameDefinitionType


@dataclass(frozen=True)
class DMLsSetBuildRequest:
    source_data_xlsx_file_path: SourceDataXlsxFilePath
    app_config: AppConfig


class DMLsSetBuilder:
    def __init__(self, a_request: DMLsSetBuildRequest):
        self.__a_request = a_request

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, a_traceback):
        """
        do nothing
        """

    def execute(self) -> DMLsSet:
        a_request = self.__a_request
        src_excel_file_path = a_request.source_data_xlsx_file_path
        app_config = a_request.app_config

        a_workbook = load_workbook(filename=src_excel_file_path.value, data_only=True)

        dmls_set = DMLsSet.empty()
        for sheet_name in a_workbook.sheetnames:
            a_worksheet = a_workbook[sheet_name]

            table_name = self.__get_table_name(a_worksheet=a_worksheet)

            db_columns_range = DMLsSetBuilder.__get_db_columns_range(
                a_worksheet=a_worksheet, app_config=app_config
            )
            # print(db_columns_range)
            db_columns_build_request = DBColumnsBuildRequest(
                source_data=db_columns_range, app_config=app_config
            )
            db_columns = DBColumnsBuilder.execute(a_request=db_columns_build_request)

            data_range = DMLsSetBuilder.__get_data_range(
                a_worksheet=a_worksheet, app_config=app_config
            )
            # print(data_range)

            dmls_build_request = DMLsBuildRequest(
                table_name=table_name, db_columns=db_columns, data_range=data_range
            )
            with DMLsBuilder(a_request=dmls_build_request) as a_dmls_builder:
                dmls = DMLs(value=a_dmls_builder.execute())

            dmls_set = dmls_set.append(key=table_name, element=dmls)

        return dmls_set

    def __get_table_name(self, a_worksheet: Worksheet) -> TableName:
        a_request = self.__a_request
        app_config = a_request.app_config

        if app_config.table_name_definition_type == TableNameDefinitionType.SHEET:
            table_name = TableName(value=a_worksheet.title)
        else:
            table_name_cell_position = app_config.table_name_cell_position
            a_cell = a_worksheet.cell(
                row=table_name_cell_position.row,
                column=table_name_cell_position.column,
            )
            table_name = TableName(value=str(a_cell.value))

        return table_name

    @staticmethod
    def __get_db_columns_range(
        a_worksheet: Worksheet, app_config: AppConfig
    ) -> list[list[str]]:
        return [
            [str(cell) for cell in col]
            for col in a_worksheet.iter_cols(
                min_row=1,
                max_row=app_config.header_max_row,
                min_col=app_config.data_start_cell_position.column,
                values_only=True,
            )
        ]

    @staticmethod
    def __get_data_range(
        a_worksheet: Worksheet, app_config: AppConfig
    ) -> list[list[str]]:
        return [
            [str(cell) for cell in row]
            for row in a_worksheet.iter_rows(
                min_row=app_config.data_start_cell_position.row,
                max_row=a_worksheet.max_row,
                min_col=app_config.data_start_cell_position.column,
                values_only=True,
            )
        ]
